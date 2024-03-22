from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Usuarios, Producto, Venta
from django.utils.timezone import now
from django.utils.dateparse import parse_date
from django.db.models import Sum
from decimal import Decimal
from .models import GastoExtra, Producto
from django.utils import timezone




# Vista existente para el índice
def index(request):
    return render(request, "index.html")

# Vistas para Usuarios
def listar(request):
    users = Usuarios.objects.all()
    datos = {'usuarios': users}
    return render(request, "crud_usuarios/listar.html", datos)

def agregar(request):
    if request.method == 'POST':
        #if request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('correo') and request.POST.get('telefono') and request.POST.get('f_nac'):
            user = Usuarios()
            user.nombre = request.POST.get('nombre')
            user.apellido = request.POST.get('apellido')
            #user.correo = request.POST.get('correo')
            #user.telefono = request.POST.get('telefono')
            #user.f_nac = request.POST.get('f_nac')
            user.save()
            return redirect('listar')
    else:
        return render(request, "crud_usuarios/agregar.html")

def actualizar(request):
    if request.method == 'POST':
        if request.POST.get('id') and request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('correo') and request.POST.get('telefono') and request.POST.get('f_nac'):
            user = Usuarios()
            user.id = request.POST.get('id')
            user.nombre = request.POST.get('nombre')
            user.apellido = request.POST.get('apellido')
            user.correo = request.POST.get('correo')
            user.telefono = request.POST.get('telefono')
            user.f_nac = request.POST.get('f_nac')
            user.f_registro = now()
            user.save()
            return redirect('listar')
    else:
        users = Usuarios.objects.all()
        datos = {'usuarios': users}
        return render(request, "crud_usuarios/actualizar.html", datos)

def eliminar(request):
    if request.method == 'POST':
        if request.POST.get('id'):
            id_borrar = request.POST.get('id')
            tupla = Usuarios.objects.get(id=id_borrar)
            tupla.delete()
            return redirect('listar')
    else:
        users = Usuarios.objects.all()
        datos = {'usuarios': users}
        return render(request, "crud_usuarios/eliminar.html", datos)

# Vistas para Productos
def listar_productos(request):
    productos = Producto.objects.all()
    total_ganancia_inv= sum(producto.ganancia_inventario for producto in productos)  # Calcular la ganancia total
    total_precio_compra_inv = sum(producto.precio_compra * producto.stock for producto in productos)
    return render(request, "productos/listar_productos.html", {'productos': productos, 'total_ganancia_inv': total_ganancia_inv, 'total_precio_compra_inv': total_precio_compra_inv })




def agregar_producto(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        precio_venta = request.POST.get('precio_venta')
        precio_compra = request.POST.get('precio_compra')
        stock = request.POST.get('stock')
        # Verificar que los campos necesarios están presentes
        if nombre and precio_venta and precio_compra and stock:
            try:
                # Convertir precio_venta y precio_compra a decimal y stock a entero
                precio_venta = float(precio_venta)
                precio_compra = float(precio_compra)
                stock = int(stock)
                
                # Crear y guardar el nuevo producto
                producto = Producto(nombre=nombre, precio_venta=precio_venta, precio_compra=precio_compra, stock=stock)
                producto.save()
                return redirect('listar_productos')
            except ValueError:
                # Manejar el caso en que la conversión de tipos falla
                # Aquí puedes decidir cómo quieres manejar el error, por ejemplo, enviando un mensaje de error al template
                pass

    # Si no es una solicitud POST o si falla la validación/conversión, renderizar el formulario nuevamente
    return render(request, "productos/agregar_producto.html")

def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        precio_venta = request.POST.get('precio_venta')
        precio_compra = request.POST.get('precio_compra')
        stock = request.POST.get('stock')
        
        if nombre and precio_venta and precio_compra and stock:
            producto.nombre = nombre
            producto.precio_venta = precio_venta
            producto.precio_compra = precio_compra
            producto.stock = stock
            producto.save()
            return redirect('listar_productos')
    return render(request, 'productos/editar_producto.html', {'producto': producto})   

def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        producto.delete()
        return redirect('listar_productos')
    return render(request, 'productos/confirmar_eliminacion.html', {'producto': producto}) 

# Vistas para Ventas
def registrar_venta(request):
    if request.method == 'POST':
        producto_id = request.POST.get('producto')
        cantidad = int(request.POST.get('cantidad'))  # Asegúrate de que la cantidad es un entero
        pagado = request.POST.get('pagado', '0') == '1'  # Captura el estado del checkbox
        observaciones = request.POST.get('observaciones', '')  # Valor por defecto cadena vacía si no se proporciona
        
        # Obtener el producto por su ID
        producto = Producto.objects.get(pk=producto_id)

        if producto and cantidad:
            # Disminuir el stock del producto
            producto.stock -= cantidad
            # Asegurarte de no tener stock negativo
            if producto.stock < 0:
                producto.stock = 0
                # Aquí podrías manejar el caso de intentar vender más producto del disponible
                # Por ejemplo, enviando un mensaje al usuario o redirigiendo a otra vista

            # Guardar el producto con su stock actualizado
            producto.save()

            # Crear y guardar la venta
            venta = Venta(producto=producto, cantidad=cantidad, pagado=pagado, observaciones=observaciones)
            venta.save()
            return redirect('listar_ventas')

    productos = Producto.objects.all()
    return render(request, "ventas/registrar_venta.html", {'productos': productos})
def listar_ventas(request):
     # Obtener parámetros de filtrado desde el request
    nombre_producto = request.GET.get('producto', None)
    fecha_inicio = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_fin', None)
    pagado = request.GET.get('pagado')
    observaciones = request.GET.get('observaciones', '')  # Nuevo filtro para observaciones


    # Filtrar queryset basado en los parámetros
    ventas = Venta.objects.all()
    
    if nombre_producto:
        ventas = ventas.filter(producto__nombre__icontains=nombre_producto)
    
    if fecha_inicio and fecha_fin:
        ventas = ventas.filter(fecha_venta__range=[fecha_inicio, fecha_fin])
    
    # Aplica filtro por estado de pago si se proporcionó
    
    if pagado == '1':
        ventas = ventas.filter(pagado=True)
    elif pagado == '0':
        ventas = ventas.filter(pagado=False)
    # Aplica filtro por observaciones si se proporcionó
    if observaciones:
        ventas = ventas.filter(observaciones__icontains=observaciones)


    # Calcular los totales basados en el queryset filtrado
    total_cantidad = ventas.aggregate(total=Sum('cantidad'))['total'] or 0
    total_venta = sum(venta.total_venta for venta in ventas)  # Asegúrate de que total_venta es calculable como propiedad
    total_ganado = sum(venta.ganancia_venta for venta in ventas)  # Asume que ganancia_venta es una propiedad en tu modelo
    


    
    # Filtrar solo ventas que están marcadas como pagadas
    ventas_pagadas = ventas.filter(pagado=True)
    # Calcular los totales basados en el queryset de ventas pagadas
    total_cantidad_pagadas = ventas_pagadas.aggregate(total=Sum('cantidad'))['total'] or 0
    total_venta_pagadas = sum(venta.total_venta for venta in ventas_pagadas)
    total_ganado_pagadas = sum(venta.ganancia_venta for venta in ventas_pagadas)

    # Calcular el 30% de las ganancias totales de ventas pagadas
    porcentaje_ganancia_pagadas = total_ganado_pagadas * Decimal('0.3')
    dinero_invertir= total_ganado_pagadas - porcentaje_ganancia_pagadas

    context = {
        'ventas': ventas,
        'total_cantidad': total_cantidad,
        'total_venta': total_venta,
        'total_ganado': total_ganado,
        # Pasa los parámetros de filtrado al template para preservarlos en el formulario
        'nombre_producto': nombre_producto,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        "pagado": pagado,
        "observaciones": observaciones,
        'total_cantidad2': total_cantidad_pagadas,
        'total_venta2': total_venta_pagadas,        
        'total_ganado2': total_ganado_pagadas,
        'porcentaje_ganancia': porcentaje_ganancia_pagadas,
         'dinero_invertir': dinero_invertir
    }

    return render(request, "ventas/listar_ventas.html", context)

from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, GastoExtra
from django.contrib import messages  # Importa esto para enviar mensajes


def actualizar_venta(request):
    if request.method == 'POST' and request.is_ajax():
        venta_id = request.POST.get('venta_id')
        pagado = bool(int(request.POST.get('pagado', 0)))
        
        try:
            venta = Venta.objects.get(pk=venta_id)
            venta.pagado = pagado
            venta.save()
            return JsonResponse({'status': 'success'})
        except Venta.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Venta no encontrada'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'})


def editar_venta(request, venta_id):
    venta = Venta.objects.get(pk=venta_id)
    if request.method == 'POST':
        # Si se envía un formulario POST, actualiza el estado de pago de la venta
        venta.pagado = not venta.pagado
        venta.save()
        return redirect('listar_ventas')  # Redirige de vuelta a la lista de ventas
    else:
        return render(request, 'ventas/editar_venta.html', {'venta': venta})


def registrar_gasto_extra(request):
    if request.method == 'POST':
        producto_id = request.POST.get('producto')
        cantidad = request.POST.get('cantidad')
        precio = request.POST.get('precio')
        descripcion = request.POST.get('descripcion', '')  # Asegúrate de proporcionar un valor por defecto
        observaciones = request.POST.get('observaciones', '')

        producto = None
        if producto_id:
            try:
                producto = Producto.objects.get(pk=producto_id)
            except Producto.DoesNotExist:
                producto = None

        # Mueve la creación de GastoExtra fuera del bloque except
        if cantidad and precio:  # Asegúrate de que cantidad y precio no estén vacíos
            try:
                cantidad = int(cantidad)
                precio = float(precio)
                gasto_extra = GastoExtra(
                    producto=producto,
                    descripcion=descripcion,
                    cantidad=cantidad,
                    precio=precio,
                    observaciones=observaciones
                )
                gasto_extra.save()
                return redirect('listar_gastos_extras')
            except ValueError:
                messages.error(request, 'Error al procesar la cantidad o el precio.')
        else:
            messages.error(request, 'La cantidad y el precio son obligatorios.')

    productos = Producto.objects.all()
    return render(request, "gastos/agregar_gasto.html", {'productos': productos})


def listar_gastos_extras(request):
    gastos_extras = GastoExtra.objects.all()
    
    # Filtrado por fecha si se proporcionan parámetros 'fecha_inicio' y 'fecha_fin'
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio and fecha_fin:
        gastos_extras = gastos_extras.filter(fecha__range=[fecha_inicio, fecha_fin])
    else:   
        gastos_extras = GastoExtra.objects.all()

   
    total_gastos = sum(gasto.cantidad * gasto.precio for gasto in gastos_extras)
    return render(request, "gastos/listar_gastos_extras.html", {'gastos_extras': gastos_extras, 'total_gastos': total_gastos})

def editar_gasto_extra(request, gasto_id):
    gasto = get_object_or_404(GastoExtra, pk=gasto_id)
    if request.method == 'POST':
        # Procesa el formulario
        # Asegúrate de validar y guardar los cambios
        return redirect('listar_gastos_extras') # Redirige a la lista de gastos extras
    else:
        # Muestra el formulario con la información del gasto para editar
        return render(request, 'gastos/editar_gasto.html', {'gasto': gasto})


def eliminar_gasto_extra(request, gasto_id):
    gasto = get_object_or_404(GastoExtra, pk=gasto_id)
    gasto.delete()
    return redirect('listar_gastos_extras')  # Asegúrate de redirigir a la lista de gastos extras

    
# views.py

from openpyxl import Workbook
from django.http import HttpResponse

############### EXPORTAR EXCEL


def exportar_excel(request):
    # Obtener los datos que quieres exportar a Excel, por ejemplo, una lista de productos
    productos = Producto.objects.all()

    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Agregar encabezados de columna
    ws.append(['ID', 'Nombre', 'Precio de venta', 'Precio de compra', 'Stock'])

    # Iterar sobre los productos y agregar cada uno a la hoja de cálculo
    for producto in productos:
        ws.append([producto.id, producto.nombre, producto.precio_venta, producto.precio_compra, producto.stock])

    # Definir el nombre del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_productos.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    wb.save(response)
    return response


def exportar_excel_ventas(request):
    # Obtener los datos que quieres exportar a Excel, por ejemplo, una lista de productos
    ventas = Venta.objects.all()

    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Agregar encabezados de columna
    ws.append(['ID', 'Producto', 'Cantidad', 'Fecha venta','Pagado','Observacion'])

    # Iterar sobre los productos y agregar cada uno a la hoja de cálculo
    for venta in ventas:
        producto_nombre = venta.producto.nombre  # Accede al atributo `nombre` del objeto Producto
        # Ajustar la fecha de venta para eliminar la información de la zona horaria
        fecha_venta = timezone.localtime(venta.fecha_venta).replace(tzinfo=None)

        ws.append([venta.id, producto_nombre, venta.cantidad, fecha_venta, venta.pagado, venta.observaciones])

    # Definir el nombre del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=listado_ventas.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    wb.save(response)
    return response


def exportar_excel_gastos(request):
    # Obtener los datos que quieres exportar a Excel, por ejemplo, una lista de productos
    gastos = GastoExtra.objects.all()

    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Agregar encabezados de columna
    ws.append(['ID', 'Fecha', 'Articulo', 'Cantidad','precio','Observacion'])

    # Iterar sobre los productos y agregar cada uno a la hoja de cálculo
    for gasto in gastos:
        # Verificar si el producto está definido antes de acceder a su nombre
        if gasto.producto is not None:
            producto_nombre = gasto.producto.nombre
        else:
            producto_nombre = gasto.descripcion  # O cualquier valor predeterminado que desees
        
        # Ajustar la fecha de venta para eliminar la información de la zona horaria
        fecha_gasto = timezone.localtime(gasto.fecha).replace(tzinfo=None)

        ws.append([gasto.id, fecha_gasto, producto_nombre, gasto.cantidad,gasto.precio, gasto.observaciones])

    # Definir el nombre del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=listado_gastos_extras.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    wb.save(response)
    return response

